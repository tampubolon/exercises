import random
import string
import os
from openpyxl import Workbook

# Function to generate a random string of a specified length
def random_string(length=5):
    return ''.join(random.choices(string.ascii_letters + string.digits, k=length))

def create_and_save_excel(rows=200000, target_files=None):
    if target_files is None:
        target_files = []

    # Create a new workbook and activate the default sheet
    wb = Workbook()
    ws = wb.active

    # Add column headers
    columns = ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J']
    for col in columns:
        ws[f'{col}1'] = col

    # Generate random data for each column and each row
    for row in range(2, rows + 2):  # Start from row 2
        for col_index in range(10):  # There are 10 columns (A to J)
            col = columns[col_index]
            if col_index % 2 == 0:  # Even index columns (A, C, E, G, I) will have integers
                ws[f'{col}{row}'] = random.randint(100, 999)  # 3-digit integers
            else:  # Odd index columns (B, D, F, H, J) will have random strings
                ws[f'{col}{row}'] = random_string(5)  # Random 5-character string

    # Save the workbook to each target file
    for file_path in target_files:
        directory = os.path.dirname(file_path)
        os.makedirs(directory, exist_ok=True)  # Ensure the directory exists
        wb.save(file_path)
        print(f"Excel file created and saved to: {file_path}")

# Example usage
create_and_save_excel(
    target_files=["task-b2/dummy-data.xlsx", "task-b1/dummy-data.xlsx"]
)
